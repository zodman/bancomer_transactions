from openpyxl import load_workbook
import tqdm
import os
os.environ["DJANGO_SETTINGS_MODULE"]="app.settings"
import django

django.setup()
import datetime
from bancomer.models import Transaction

wb = load_workbook(filename="descargas.xlsx")
ws = wb.active
DATE_FORMAT="%d/%m/%Y"


for i in tqdm.trange(5, 1000):
    for j in ws[f"A{i}":f"D{i}"]:
        datestr, desc, saldo, abono = [k.value for k in j ]
        if desc is None and saldo is None and abono is None:
            break
        
        if not saldo  and abono:
            amount = abono
        else:
            amount = saldo
        date_obj = datetime.datetime.strptime(datestr, DATE_FORMAT).date()
        tqdm.tqdm.write("{} {} [{}]".format(date_obj, desc, amount) )
        Transaction.objects.get_or_create(applied=date_obj, description=desc,amount=amount)  
